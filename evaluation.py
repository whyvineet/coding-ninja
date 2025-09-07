"""
Enhanced Evaluation system for Excel mock interviewer
Handles both text answers and Excel file uploads with comprehensive scoring
"""

import pandas as pd
from openpyxl import load_workbook
import io
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain_core.prompts import ChatPromptTemplate
from langchain_core.output_parsers import JsonOutputParser
from typing import Dict, Any, List, Optional
from dataclasses import dataclass
import logging
import re
from datetime import datetime


@dataclass
class EvaluationResult:
    """Structured evaluation result"""
    score: float  # 0-10 scale
    feedback: str  # Constructive feedback
    strengths: List[str]  # What they did well
    improvements: List[str]  # Areas for improvement
    category_scores: Optional[Dict[str, float]] = None  # Breakdown by skill area


class AnswerEvaluator:
    """Enhanced evaluator with better error handling and scoring"""
    
    def __init__(self, gemini_api_key: str):
        self.llm = ChatGoogleGenerativeAI(
            model="gemini-1.5-pro", 
            google_api_key=gemini_api_key,
            temperature=0.2
        )
        
        # Enhanced scoring rubrics
        self.text_rubric = {
            "technical_accuracy": {"weight": 0.35, "description": "Correctness of Excel concepts and methods"},
            "practical_application": {"weight": 0.25, "description": "Real-world applicability and examples"},
            "clarity_communication": {"weight": 0.25, "description": "Clear explanation and structure"},
            "completeness": {"weight": 0.15, "description": "Addresses all aspects of the question"}
        }
        
        self.excel_rubric = {
            "formula_correctness": {"weight": 0.40, "description": "Correct formulas and calculations"},
            "data_structure": {"weight": 0.25, "description": "Proper data organization and layout"},
            "functionality": {"weight": 0.20, "description": "Working solution that meets requirements"},
            "best_practices": {"weight": 0.15, "description": "Excel best practices and efficiency"}
        }
        
        # Set up logging
        logging.basicConfig(level=logging.INFO)
        self.logger = logging.getLogger(__name__)
    
    def evaluate_text_answer(self, answer: str, question_data: Dict[str, Any]) -> EvaluationResult:
        """Enhanced text evaluation with robust error handling and retry logic"""
        
        max_retries = 3
        
        for attempt in range(max_retries):
            try:
                prompt = f"""You are an expert Excel interviewer with 15+ years of experience evaluating candidate responses.

QUESTION CONTEXT:
Question: {question_data.get('question_text', '')}
Skill Area: {question_data.get('skill_area', 'General Excel')}
Difficulty Level: {question_data.get('difficulty_level', 2)}/5
Question Type: {question_data.get('question_type', 'text')}

CANDIDATE'S ANSWER:
{answer}

EVALUATION CRITERIA:
Please evaluate based on these weighted criteria:
1. Technical Accuracy (35%): Correctness of Excel concepts, functions, and methodology
2. Practical Application (25%): Real-world relevance, examples, use cases
3. Clarity & Communication (25%): Clear explanation, logical structure, terminology
4. Completeness (15%): Addresses all parts of the question thoroughly

SCORING SCALE:
9-10: Exceptional - Expert-level knowledge with nuanced understanding
7-8: Strong - Good grasp with minor gaps or improvements possible
5-6: Adequate - Basic understanding but missing key elements
3-4: Below Average - Significant gaps in knowledge or application
1-2: Poor - Major misconceptions or very limited understanding

CRITICAL: Return ONLY valid JSON with ALL required fields:
{{{{
    "overall_score": 0-10,
    "category_scores": {{{{
        "technical_accuracy": 0-10,
        "practical_application": 0-10,
        "clarity_communication": 0-10,
        "completeness": 0-10
    }}}},
    "strengths": ["specific strength 1", "specific strength 2"],
    "improvements": ["specific improvement 1", "specific improvement 2"],
    "detailed_feedback": "2-3 sentences of constructive feedback focusing on what they did well and how to improve"
}}}}

Ensure the JSON is valid and complete."""

                prompt_template = ChatPromptTemplate.from_template(prompt)
                parser = JsonOutputParser()
                chain = prompt_template | self.llm | parser
                eval_data = chain.invoke({})
                
                # Robust validation
                self._validate_evaluation_response(eval_data)
                
                # Ensure all required fields have proper values
                eval_data.setdefault("strengths", ["Shows understanding of the topic"])
                eval_data.setdefault("improvements", ["Consider providing more specific details"])
                eval_data.setdefault("category_scores", {})
                
                # Ensure score is within valid range
                score = float(eval_data["overall_score"])
                score = max(0.0, min(10.0, score))
                
                return EvaluationResult(
                    score=score,
                    feedback=eval_data["detailed_feedback"],
                    strengths=eval_data.get("strengths", []),
                    improvements=eval_data.get("improvements", []),
                    category_scores=eval_data.get("category_scores", {})
                )
                
            except Exception as e:
                self.logger.warning(f"Text evaluation attempt {attempt + 1} failed: {e}")
                if attempt == max_retries - 1:
                    # Final attempt - create basic evaluation
                    return self._create_basic_text_evaluation(answer, question_data)
                
                # Wait before retrying
                import time
                time.sleep(1 * (attempt + 1))
    
    def _create_basic_text_evaluation(self, answer: str, question_data: Dict[str, Any]) -> EvaluationResult:
        """Create a basic evaluation when Gemini fails"""
        
        # Basic scoring based on answer characteristics
        word_count = len(answer.split())
        score = 5.0  # Base score
        
        # Adjust based on length
        if word_count >= 50:
            score += 1.0
        elif word_count >= 25:
            score += 0.5
        elif word_count < 10:
            score -= 2.0
        
        # Check for Excel terminology
        excel_terms = ['excel', 'formula', 'function', 'cell', 'range', 'worksheet', 'data', 'chart']
        found_terms = [term for term in excel_terms if term.lower() in answer.lower()]
        score += min(2.0, len(found_terms) * 0.3)
        
        # Ensure score is within bounds
        score = max(1.0, min(10.0, score))
        
        return EvaluationResult(
            score=round(score, 1),
            feedback="Basic evaluation completed. For detailed feedback, please ensure stable connection and try again.",
            strengths=["Provided a response to the question"],
            improvements=["Consider providing more detailed explanations with specific Excel examples"]
        )
    
    def evaluate_excel_upload(self, file_content: bytes, task_description: str) -> EvaluationResult:
        """Enhanced Excel file evaluation with robust error handling"""
        
        try:
            # Load and analyze Excel file
            analysis = self._comprehensive_excel_analysis(file_content)
            
            # Use LLM for intelligent evaluation with retry logic
            max_retries = 3
            
            for attempt in range(max_retries):
                try:
                    return self._llm_excel_evaluation(analysis, task_description)
                except Exception as e:
                    self.logger.warning(f"Excel LLM evaluation attempt {attempt + 1} failed: {e}")
                    if attempt == max_retries - 1:
                        # Final attempt failed - use analysis-based evaluation
                        return self._analysis_based_evaluation(analysis, task_description)
                    
                    # Wait before retrying
                    import time
                    time.sleep(1 * (attempt + 1))
            
        except Exception as e:
            self.logger.error(f"Excel evaluation error: {e}")
            return EvaluationResult(
                score=2.0,
                feedback=f"Unable to process Excel file: {str(e)}. Please ensure the file is a valid Excel format (.xlsx or .xls) and try again.",
                strengths=[],
                improvements=["Verify file format and structure", "Ensure file is not corrupted", "Try uploading the file again"]
            )
    
    def _comprehensive_excel_analysis(self, file_content: bytes) -> Dict[str, Any]:
        """Enhanced Excel file analysis"""
        
        analysis = {
            "file_readable": False,
            "worksheets_count": 0,
            "data_present": False,
            "formulas": [],
            "charts_count": 0,
            "pivot_tables": 0,
            "functions_used": set(),
            "data_summary": {},
            "formatting_features": [],
            "errors": []
        }
        
        try:
            # Load workbook
            workbook = load_workbook(io.BytesIO(file_content), data_only=False)
            analysis["file_readable"] = True
            analysis["worksheets_count"] = len(workbook.worksheets)
            
            worksheet = workbook.active
            
            # Analyze formulas and functions
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula = cell.value
                        analysis["formulas"].append({
                            "cell": cell.coordinate,
                            "formula": formula
                        })
                        
                        # Extract functions used
                        functions = re.findall(r'([A-Z]+)\(', formula.upper())
                        analysis["functions_used"].update(functions)
            
            # Charts analysis
            analysis["charts_count"] = len(worksheet._charts)
            
            # Data analysis using pandas
            try:
                df = pd.read_excel(io.BytesIO(file_content))
                if not df.empty:
                    analysis["data_present"] = True
                    analysis["data_summary"] = {
                        "rows": len(df),
                        "columns": len(df.columns),
                        "column_names": df.columns.tolist()[:10],  # First 10 columns
                        "numeric_columns": df.select_dtypes(include=['number']).columns.tolist(),
                        "has_headers": True if df.columns.dtype == 'object' else False
                    }
            except Exception as e:
                analysis["errors"].append(f"Data reading error: {str(e)}")
                
        except Exception as e:
            analysis["errors"].append(f"File analysis error: {str(e)}")
            
        # Convert set to list for JSON serialization
        analysis["functions_used"] = list(analysis["functions_used"])
        
        return analysis
    
    def _llm_excel_evaluation(self, analysis: Dict[str, Any], task_description: str) -> EvaluationResult:
        """Use Gemini to evaluate Excel file based on analysis with robust error handling"""
        
        prompt = f"""You are an expert Excel instructor evaluating a student's Excel file submission.

TASK DESCRIPTION:
{task_description}

FILE ANALYSIS RESULTS:
- File readable: {analysis['file_readable']}
- Number of worksheets: {analysis['worksheets_count']}
- Data present: {analysis['data_present']}
- Number of formulas: {len(analysis['formulas'])}
- Functions used: {', '.join(analysis['functions_used']) if analysis['functions_used'] else 'None'}
- Charts: {analysis['charts_count']}
- Data summary: {analysis['data_summary']}
- Formulas found: {[f["formula"] for f in analysis['formulas'][:5]]}  # First 5 formulas
- Errors: {analysis['errors']}

EVALUATION CRITERIA:
1. Formula Correctness (40%): Are the formulas accurate and appropriate?
2. Data Structure (25%): Is the data well-organized and properly formatted?
3. Functionality (20%): Does the solution work and meet requirements?
4. Best Practices (15%): Does it follow Excel best practices?

SCORING SCALE:
9-10: Exceptional work, exceeds expectations
7-8: Good work with minor issues
5-6: Adequate but missing key elements
3-4: Below average with significant issues
1-2: Poor work with major problems

CRITICAL: Return ONLY valid JSON with ALL required fields:
{{{{
    "overall_score": 0-10,
    "category_scores": {{{{
        "formula_correctness": 0-10,
        "data_structure": 0-10,
        "functionality": 0-10,
        "best_practices": 0-10
    }}}},
    "strengths": ["specific strength 1", "specific strength 2"],
    "improvements": ["specific improvement 1", "specific improvement 2"],
    "detailed_feedback": "2-3 sentences of constructive feedback"
}}}}

Ensure the JSON is valid and complete."""

        prompt_template = ChatPromptTemplate.from_template(prompt)
        parser = JsonOutputParser()
        chain = prompt_template | self.llm | parser
        eval_data = chain.invoke({})
        
        # Robust validation
        self._validate_evaluation_response(eval_data)
        
        # Ensure all required fields have proper values
        eval_data.setdefault("strengths", ["File uploaded successfully"])
        eval_data.setdefault("improvements", ["Consider adding more comprehensive documentation"])
        eval_data.setdefault("category_scores", {})
        
        # Ensure score is within valid range
        score = float(eval_data["overall_score"])
        score = max(0.0, min(10.0, score))
        
        return EvaluationResult(
            score=score,
            feedback=eval_data["detailed_feedback"],
            strengths=eval_data.get("strengths", []),
            improvements=eval_data.get("improvements", []),
            category_scores=eval_data.get("category_scores", {})
        )
    
    def _analysis_based_evaluation(self, analysis: Dict[str, Any], task_description: str) -> EvaluationResult:
        """Generate evaluation based on file analysis when LLM fails"""
        
        score = 5.0  # Base score
        feedback_parts = []
        strengths = []
        improvements = []
        
        # File readability check
        if not analysis["file_readable"]:
            score = 2.0
            feedback_parts.append("File could not be read properly.")
            improvements.append("Ensure file is in correct Excel format (.xlsx or .xls)")
        else:
            strengths.append("File is properly formatted and readable")
            
        # Data presence check
        if analysis["data_present"]:
            score += 1.0
            strengths.append("Contains data as expected")
        else:
            score -= 1.0
            improvements.append("Include relevant data in the spreadsheet")
            
        # Formula analysis
        if analysis["formulas"]:
            score += 1.5
            strengths.append(f"Uses {len(analysis['formulas'])} formulas")
            
            # Check for advanced functions
            advanced_functions = {'VLOOKUP', 'INDEX', 'MATCH', 'SUMIF', 'COUNTIF', 'PIVOT'}
            used_advanced = set(analysis["functions_used"]) & advanced_functions
            if used_advanced:
                score += 1.0
                strengths.append(f"Uses advanced functions: {', '.join(used_advanced)}")
        else:
            score -= 1.0
            improvements.append("Include relevant formulas to solve the problem")
            
        # Charts and visualization
        if analysis["charts_count"] > 0:
            score += 0.5
            strengths.append(f"Includes {analysis['charts_count']} chart(s)")
            
        # Error handling
        if analysis["errors"]:
            score -= 0.5
            improvements.append("Address file structure issues")
            
        score = max(1.0, min(10.0, score))
        
        feedback = f"Analysis-based evaluation: {' '.join(feedback_parts) if feedback_parts else 'Basic Excel file analysis completed.'}"
        
        return EvaluationResult(
            score=round(score, 1),
            feedback=feedback,
            strengths=strengths,
            improvements=improvements
        )
    
    def _validate_evaluation_response(self, eval_data: Dict[str, Any]) -> None:
        """Validate Gemini evaluation response with enhanced checks"""
        required_fields = ["overall_score", "detailed_feedback"]
        
        for field in required_fields:
            if field not in eval_data:
                raise ValueError(f"Missing required field: {field}")
                
        score = eval_data["overall_score"]
        if not isinstance(score, (int, float)) or not (0 <= score <= 10):
            raise ValueError(f"Invalid score: {score}")
            
        # Ensure feedback is meaningful
        feedback = eval_data.get("detailed_feedback", "")
        if not feedback or len(feedback.strip()) < 10:
            raise ValueError("Feedback too short or empty")
            
        # Ensure lists are properly formatted
        for list_field in ["strengths", "improvements"]:
            if list_field in eval_data:
                if not isinstance(eval_data[list_field], list):
                    raise ValueError(f"{list_field} must be a list")


def create_performance_report(interview_data: Dict[str, Any]) -> Dict[str, Any]:
    """Create a comprehensive performance report from interview data"""
    
    if not interview_data.get('question_history'):
        return {
            "error": "No interview data available",
            "report": {}
        }
    
    questions = interview_data['question_history']
    total_questions = len(questions)
    
    # Calculate overall metrics
    scores = [q.get('score', 0) for q in questions if q.get('score') is not None]
    overall_score = sum(scores) / len(scores) if scores else 0
    
    # Skill area breakdown
    skill_breakdown = {}
    for question in questions:
        skill = question.get('skill_area', 'General Excel')
        if skill not in skill_breakdown:
            skill_breakdown[skill] = {'scores': [], 'questions': 0}
        
        skill_breakdown[skill]['questions'] += 1
        if question.get('score') is not None:
            skill_breakdown[skill]['scores'].append(question['score'])
    
    # Calculate averages for each skill
    for skill_data in skill_breakdown.values():
        scores = skill_data['scores']
        skill_data['average_score'] = sum(scores) / len(scores) if scores else 0
        skill_data['performance_level'] = _get_performance_level(skill_data['average_score'])
    
    # Collect all feedback
    all_strengths = []
    all_improvements = []
    for question in questions:
        all_strengths.extend(question.get('strengths', []))
        all_improvements.extend(question.get('improvements', []))
    
    # Performance level
    performance_level = _get_performance_level(overall_score)
    
    # Generate recommendations
    recommendations = _generate_recommendations(overall_score, skill_breakdown, all_improvements)
    
    report = {
        "candidate_name": interview_data.get('candidate_name', 'Unknown'),
        "session_id": interview_data.get('session_id', ''),
        "interview_date": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        "overall_score": round(overall_score, 1),
        "performance_level": performance_level,
        "total_questions": total_questions,
        "questions_completed": len([q for q in questions if q.get('score') is not None]),
        "skill_breakdown": {
            skill: {
                "average_score": round(data['average_score'], 1),
                "questions_asked": data['questions'],
                "performance_level": data['performance_level']
            }
            for skill, data in skill_breakdown.items()
        },
        "strengths": list(set(all_strengths)),  # Remove duplicates
        "improvements": list(set(all_improvements)),  # Remove duplicates
        "recommendations": recommendations,
        "detailed_questions": [
            {
                "question_number": i + 1,
                "question_text": q.get('question_text', ''),
                "skill_area": q.get('skill_area', ''),
                "difficulty_level": q.get('difficulty_level', 1),
                "score": q.get('score'),
                "feedback": q.get('feedback', ''),
                "strengths": q.get('strengths', []),
                "improvements": q.get('improvements', [])
            }
            for i, q in enumerate(questions)
        ]
    }
    
    return {"report": report, "error": None}


def _get_performance_level(score: float) -> str:
    """Get performance level description based on score"""
    if score >= 8.5:
        return "Expert"
    elif score >= 7.0:
        return "Advanced"
    elif score >= 5.0:
        return "Intermediate"
    elif score >= 3.0:
        return "Beginner"
    else:
        return "Needs Improvement"


def _generate_recommendations(overall_score: float, skill_breakdown: Dict, improvements: List[str]) -> List[str]:
    """Generate personalized learning recommendations"""
    recommendations = []
    
    # Overall recommendations based on score
    if overall_score < 5.0:
        recommendations.append("Start with Excel basics: cell references, simple formulas, and basic functions")
        recommendations.append("Practice using SUM, AVERAGE, COUNT functions regularly")
    elif overall_score < 7.0:
        recommendations.append("Focus on intermediate Excel features like VLOOKUP and pivot tables")
        recommendations.append("Learn conditional formatting and data validation")
    else:
        recommendations.append("Explore advanced Excel features like Power Query and Power Pivot")
        recommendations.append("Consider learning VBA for automation")
    
    # Skill-specific recommendations
    weak_skills = [skill for skill, data in skill_breakdown.items() 
                   if data['average_score'] < overall_score - 1.0]
    
    for skill in weak_skills:
        if 'lookup' in skill.lower():
            recommendations.append("Practice VLOOKUP, INDEX/MATCH functions with different datasets")
        elif 'pivot' in skill.lower():
            recommendations.append("Create pivot tables with various data sources and analyze trends")
        elif 'chart' in skill.lower():
            recommendations.append("Learn different chart types and when to use each effectively")
        elif 'formula' in skill.lower():
            recommendations.append("Study array formulas and nested function combinations")
    
    # Add improvement-based recommendations
    common_improvements = [imp for imp in improvements if improvements.count(imp) > 1]
    for improvement in common_improvements[:3]:  # Top 3 most common
        if 'example' in improvement.lower():
            recommendations.append("Practice explaining Excel concepts with real-world examples")
        elif 'detail' in improvement.lower():
            recommendations.append("Work on providing more comprehensive explanations")
    
    return list(set(recommendations))  # Remove duplicates